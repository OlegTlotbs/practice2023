import requests
import json
import os
import sys
import io
import pandas as pd
import openpyxl
from itertools import zip_longest
from bs4 import BeautifulSoup as bs
from collections import OrderedDict 



# -----------------------------------------------



def createId(link):
    if "https://catalog.data.gov/dataset/" not in link:
        return "temp"
    return link.replace("https://catalog.data.gov/dataset/", "")

def getReadableSize(size):
    power = 2**10  # 1024
    n = 0
    units = {0: '', 1: 'КБ', 2: 'МБ', 3: 'ГБ', 4: 'ТБ'}
    while size >= power:
        size /= power
        n += 1
    size = round(size, 2)
    return f"{size}{units[n]}"

def findTitle(soup):
    title = soup.find("h1", itemprop="name")
    if title is None:
        return None
    return title.text.strip()

def findDataInTable(soup, data):
    header = soup.find("th", string = data)
    if header is not None:
        row = header.parent
        value = row.find("td")
        if value is not None:
            return value.text.strip()
    return None

def getMediaTypeVocabulary():
    fileName = "MediaTypeVocabulary.txt"
    
    if os.path.isfile(fileName):
        with open(fileName, "r", encoding="utf-8") as file:
            mediaTypeVocabulary = [line.strip() for line in file.readlines()]
            return mediaTypeVocabulary
    
    url = "https://www.iana.org/assignments/media-types/media-types.xhtml"
    r = requests.get(url)
    soup = bs(r.text, "html.parser")
    
    tableT = soup.find_all("th", string = "Template")
    
    mediaTypeVocabulary = []
    
    for tt in tableT:
        table = tt.parent.parent.parent
        column2 = table.find_all("td")[1::3]
        for c2 in column2:
            tag = c2.find("a")
            if tag is not None:
                mediaTypeVocabulary.append(tag["href"])
    
    with open(fileName, "w", encoding="utf-8") as file:
        for mt in mediaTypeVocabulary:
            file.write(mt + "\n")

    return mediaTypeVocabulary

def getLicencesVocabulary():
    fileName = "LicencesVocabulary.txt"
    
    if os.path.isfile(fileName):
        with open(fileName, "r", encoding="utf-8") as file:
            licencesVocabulary = [line.strip() for line in file.readlines()]
            return licencesVocabulary
    
    url = "https://gitlab.com/european-data-portal/edp-vocabularies/-/raw/master/edp-licences-skos.rdf?inline=false"
    r = requests.get(url)
    soup = bs(r.text, "xml")
    
    licencesVocabulary = []
    
    tags = soup.find_all("dc:identifier")
    for tag in tags:
        licencesVocabulary.append(tag.text)
    tags = soup.find_all("skos:prefLabel")
    for tag in tags:
        licencesVocabulary.append(tag.text)
    
    with open(fileName, "w", encoding="utf-8") as file:
        for l in licencesVocabulary:
            file.write(l + "\n")

    return licencesVocabulary



# -----------------------------------------------



def findFormats(soup):
    formats = []
    downloadSection = soup.find("section", id = "dataset-resources")
    if downloadSection is not None:
        formatSpans = downloadSection.find_all("span", class_ = "format-label")
        for format in formatSpans:
            formats.append(format.text)
        formats = list(OrderedDict.fromkeys(formats))
    
    print(" @ formats: " + str(formats)) # ////
    
    formats = [
        f for f in formats if "Landing Page" != f and "Esri REST" != f
    ]
    
    return formats

def haveFormats(formats):
    if not formats:
        return False
    for f in formats:
        if f != "":
            return True
    return False



def findSourceFormats(soup):
    formats = []
    downloadSection = soup.find("section", id = "dataset-metadata-source")
    if downloadSection is not None:
        formatSpans = downloadSection.find_all("span", class_ = "format-label")
        for format in formatSpans:
            formats.append(format.text)
        formats = list(OrderedDict.fromkeys(formats))
    print(" @ source: " + str(formats)) # ////
    return formats

def findSourceUrl(soup):
    downloadSection = soup.find("section", id = "dataset-metadata-source")
    url = None
    if downloadSection is not None:
        link = downloadSection.find("a", string = "Download Metadata")
        link = link["href"]
        url = "https://catalog.data.gov" + link
    print(" @ source url: " + str(url)) # ////
    return url

def findMediaType(soup):
    mediaDownloadURL = []
    
    formatsSource = findSourceFormats(soup)
    if "Data.json" not in formatsSource:
        return [], mediaDownloadURL
    
    url = findSourceUrl(soup)
    if url is None:
        return [], mediaDownloadURL
    
    r = requests.get(url)
    if r.status_code != 200:
        return [], mediaDownloadURL

    mediaTypes = []

    jd = json.loads(r.text)
    if "distribution" in jd:
        distr = jd["distribution"]
        for d in distr:
            if "mediaType" in d:
                mediaTypes.append(d["mediaType"])
        for d in distr:
            if "mediaType" in d and "downloadURL" in d:
                mediaDownloadURL.append((d["mediaType"], d["downloadURL"]))
    
    return mediaTypes, mediaDownloadURL

def haveMediaTypes(mediaTypes):
    if not mediaTypes:
        return False
    for mt in mediaTypes:
        if mt != "":
            return True
    return False



def isVocabularyMediaType(mediaTypes, mediaTypeVocabulary):
    if not mediaTypes:
        print(" @ нет медиа типов") # ////
        return False
    for mt in mediaTypes:
        if mt not in mediaTypeVocabulary:
            print(" @ нет медиа типа: " + mt) # ////
            return False
    return True



def isNonProprietaryFormat(formats):
    if not formats:
        return False

    nonProprietaryFormats = ["BMP","CSV","DBF","GEOJSON","GZIP","HTML","ICS","JPEG2000","JSON","JSON_LD","KML","KMZ","NETCDF","ODS","PNG","RDF","RDF_N_QUADS","RDF_N_TRIPLES","RDF_TRIG","RDF_TURTLE","RDF_XML","RSS","RTF","TAR","TIFF","TSV","TXT","WMS_SRVC","XML","ZIP"]
    
    for f in formats:
        if f not in nonProprietaryFormats:
            return False
    return True

def isMachineReadableFormats(formats):
    if not formats:
        return False

    machineReadableFormats = ["CSV","GEOJSON","ICS","JSON","JSON_LD","KML","KMZ","NETCDF","ODS","RDF","RDFA","RDF_N_QUADS","RDF_N_TRIPLES","RDF_TRIG","RDF_TURTLE","RDF_XML","RSS","SHP","XLS","XLSX","XML"]
    
    for f in formats:
        if f not in machineReadableFormats:
            return False
    return True



def downloadData(link, id):
    url = link[1]
    
    response = requests.head(url)
    content_type = response.headers.get("content-type")
    if content_type and content_type.startswith("text/html"):
        print(" @ по ссылке на скачивание нет файла:", link[0]) # ////
        return None
    
    filePath = os.path.join("temp", id + "_" + link[0])
    filePath = filePath.replace("/", "_")
    filePath = filePath.replace(".", "_")
    
    if os.path.isfile(filePath):
        print(" @ файл найден на диске:", link[0]) # ////
        with open(filePath, "r", encoding="utf-8") as file:
            data = file.read()
            return data
    
    dataReq = requests.get(url)
        
    if dataReq.status_code != 200:
        print(" @ не удалось скачать файл:", link[0]) # ////
        return None
    else:
        data = dataReq.text
        size = "[" + getReadableSize(sys.getsizeof(data)) + "]"
        print(" @ файл скачен:", link[0], size) # ////
        
        if not os.path.exists("temp"):
            os.mkdir("temp")
        with open(filePath, "w", encoding="utf-8") as file:
            file.write(data)
        
        return data

def checkComplianceDCATAP(mediaDownloadURL, id):
    print(" @ ---DCAT-AP---") # ////
    r = checkComplianceDCATAP_p(mediaDownloadURL, id)
    print(" @ ---DCAT-AP---") # ////
    return r

def checkComplianceDCATAP_p(mediaDownloadURL, id):
    if not mediaDownloadURL:
        print(" @ нет медиа-типа или ссылки скачивания") # ////
        return False
    
    for md in mediaDownloadURL:
        url = md[1]
        
        data = downloadData(md, id)
        
        if data:
            size = "[" + getReadableSize(sys.getsizeof(data)) + "]"
            headers = { "Content-Type": md[0] }
            
            serviseUrl = "https://data.europa.eu/api/mqa/shacl/validation/report"
            try:
                response = requests.post(serviseUrl, headers = headers, data = data)
                if response.status_code == 200:
                    print(" @ проверка успешна, файл:     ", md[0], size) # ////
                elif response.status_code == 400:
                    print(" @ проверка провалилась, файл: ", md[0], size) # ////
                    print(" @  : ", response.text)
                    return False
                else:
                    print(" @ произошла ошибка1, файл:    ", md[0], size) # ////
                    print(" @  : ", response.text)
                    return False
            except requests.exceptions.RequestException as e:
                print(" @ произошла ошибка2, файл:        ", md[0], size) # ////
                return False
        else:
            return False
    return True



# -----------------------------------------------



def findLicense(soup):
    licenses = []

    l = findDataInTable(soup, "License")
    if l is not None:
        licenses.append(l)
    print(" @ License:     " + str(l)) # ////

    l = findDataInTable(soup, "License Url")
    if l is not None:
        licenses.append(l)
    print(" @ License Url: " + str(l)) # ////

    l = findDataInTable(soup, "Licence")
    if l is not None:
        licenses.append(l)
    print(" @ Licence:     " + str(l)) # ////
    
    l = findDataInTable(soup, "Licence Url")
    if l is not None:
        licenses.append(l)
    print(" @ Licence Url: " + str(l)) # ////
    
    l = None
    licenseTag = soup.find("strong", string="License:")
    if licenseTag is not None:
        licenseSpan = licenseTag.parent
        l = licenseSpan.text.strip()
        if l.startswith("License:"):
            l = l[8:]
            l = l.strip()
        
        licenses.append(l)
    print(" @ License Tag: " + str(l)) # ////
    
    licenses = [
        s for s in licenses if "No license information was provided" not in s
    ]
    
    licenses = list(set(licenses))
    return licenses

def haveLicense(licenses):
    return bool(licenses)



def isVocabularyLicense(licenses, licencesVocabulary):
    if not licenses:
        return False
    
    for l in licenses:
        if l in licencesVocabulary:
            return True
    return False



def findAccessRestrictions(soup):
    return findDataInTable(soup, "Public Access Level")

def haveAccessRestrictions(access):
    return access is not None

def isAccessRestrictionsVocabulary(access):
    vocabulary = [
        "confidential",
        "non-public",
        "public",
        "restricted",
        "sensitive"
    ]
    
    return access in vocabulary



def isNoreply(email):
    if "no-reply" not in email:
        return False
    else:
        return True

def haveContact(soup):
    contactTag = soup.find("a", title="contact")
    if contactTag is None:
        print(" @ нет тега контакта") # ////
        return False
    email = contactTag["href"]
    print(" @ mail: " + email) # ////
    if isNoreply(email):
        return False
    else:
        return True



def havePublisher(soup):
    publsherTag = soup.find("a", title="publsher")
    if publsherTag is None:
        print(" @ нет тега издателя") # ////
        return False
    print(" @ publisher: " + publsherTag.text) # ////
    return True



# -----------------------------------------------



def findDownloadLinks(soup):
    links = []
    
    downloadSection = soup.find("section", id = "dataset-resources")
    if downloadSection is not None:
        buttons = downloadSection.find_all("i", class_ = "fa fa-download")
        for b in buttons:
            l = b.parent
            links.append((l["data-format"], l["href"]))
            
    return links

def downloadDataFile(link, id):
    url = link[1]
    
    response = requests.head(url)
    content_type = response.headers.get("content-type")
    if content_type and content_type.startswith("text/html"):
        print(" @ по ссылке на скачивание нет файла:", link[0]) # ////
        return None
    
    filePath = os.path.join("temp", id + "_" + link[0])
    
    if os.path.isfile(filePath):
        print(" @ файл найден на диске:", link[0]) # ////
        return filePath
    
    dataReq = requests.get(url)
        
    if dataReq.status_code != 200:
        print(" @ не удалось скачать файл:", link[0]) # ////
        return None
    else:
        data = dataReq.text
        size = "[" + getReadableSize(sys.getsizeof(data)) + "]"
        print(" @ файл скачен:", link[0], size) # ////
        
        if not os.path.exists("temp"):
            os.mkdir("temp")
        with open(filePath, "w", encoding="utf-8") as file:
            file.write(data)
        
        return filePath

def checkFiles(downloadLinks, id):
    print(" @ ---Files---") # ////
    r = checkFiles_d(downloadLinks, id)
    print(" @ ---Files---") # ////
    return r
    
def checkFiles_d(downloadLinks, id):
    if not downloadLinks:
        print(" @ нет файлов для скачивания") # ////
        return None
    
    data = None
    for dl in downloadLinks:
        if dl[0] == "csv":
            file = downloadDataFile(dl, id)
            if file is not None:
                data = pd.read_csv(file)
                if not isinstance(data, pd.DataFrame) or data.empty:
                    print(" @ не удалось прочитать файл: ", dl[0]) # ////
                    data = None
                else:
                    break
                        
    if data is None:
        print(" @ не удалось найти подходящего формата") # ////
        return None
    
    num_rows = data.shape[0]
    num_columns = data.shape[1]
    
    unique_values = list(data.nunique().items())
    missing_values = list(data.isnull().sum().items())
    
    unique_values = [x[1] for x in unique_values]
    missing_values = [x[1] for x in missing_values]
    
    
    dataName = list(data)
    dataNumbersName = list(data.select_dtypes(include=["float64", "int64"]))
    
    column_names = dataName
    amount_zero = []
    min_values = []
    max_values = []
    mean_values = []
    
    for i in range(data.shape[1]):
        if dataName[i] in dataNumbersName:
            col = data[dataName[i]]
            
            amount_zero.append(col[col == 0].count())
            min_values.append(col.min())
            max_values.append(col.max())
            mean_values.append(round(col.mean(), 2))
        else:
            amount_zero.append("-")
            min_values.append("-")
            max_values.append("-")
            mean_values.append("-")

    res = {
        "num_rows":       num_rows,
        "num_columns":    num_columns,
        "column_names":   column_names,
        "unique_values":  unique_values,
        "missing_values": missing_values,
        "amount_zero":    amount_zero,
        
        "min_values":     min_values,
        "max_values":     max_values,
        "mean_values":    mean_values,
    }
    
    return res

def printInfo(r):
    if r is None:
        print("нет табличных данных")
    else:
        num_columns = r["num_columns"]
        unique_values = ["Уникальные"] + r["unique_values"]
        missing_values = ["Пропуски"] + r["missing_values"]
        amount_zero = ["Нули"] + r["amount_zero"]
        
        min_values = ["Мин"] + r["min_values"]
        max_values = ["Макс"] + r["max_values"]
        mean_values = ["Сред"] + r["mean_values"]
        
        print("количество строк:   ", r["num_rows"])
        print("количество столбцов:", num_columns)
        
        print("---")
        
        randeValues = ["№"] + list(range(1, num_columns+1))
        
        data = zip_longest(randeValues, unique_values, missing_values, amount_zero,
                           min_values, max_values, mean_values, fillvalue='')

        for values in data:
            print("{:<5} {:<12} {:<12} {:<12} {:<15} {:<15} {:<15}".format(*values))



# -----------------------------------------------



def printConsole(Interoperability_Info, Reusability_Info, File_Info):
    Format                            = Interoperability_Info["Format"]
    Media_type                        = Interoperability_Info["Media_type"]
    Format_Media_type_from_vocabulary = Interoperability_Info["Format_Media_type_from_vocabulary"]
    Non_proprietary                   = Interoperability_Info["Non_proprietary"]
    Machine_readable                  = Interoperability_Info["Machine_readable"]
    DCATAP_compliance                 = Interoperability_Info["DCATAP_compliance"]
    interoperabilityPoints            = Interoperability_Info["InteroperabilityPoints"]

    License_information            = Reusability_Info["License_information"]
    License_vocabulary             = Reusability_Info["License_vocabulary"]
    Access_restrictions            = Reusability_Info["Access_restrictions"]
    Access_restrictions_vocabulary = Reusability_Info["Access_restrictions_vocabulary"]
    Contact_point                  = Reusability_Info["Contact_point"]
    Publisher                      = Reusability_Info["Publisher"]
    reusabilityPoints              = Reusability_Info["ReusabilityPoints"]
    
    print("---Interoperability---")
    print("[20] Format:                              " + str(Format))
    print("[10] Media type:                          " + str(Media_type))
    print("[10] Format / Media type from vocabulary: " + str(Format_Media_type_from_vocabulary))
    print("[20] Non-proprietary:                     " + str(Non_proprietary))
    print("[20] Machine readable:                    " + str(Machine_readable))
    print("[30] DCAT-AP compliance:                  " + str(DCATAP_compliance))
    print("Rating Interoperability:                  " + str(interoperabilityPoints))
    print("---Reusability---")
    print("[20] License information:                 " + str(License_information))
    print("[10] License vocabulary:                  " + str(License_vocabulary))
    print("[10] Access restrictions:                 " + str(Access_restrictions))
    print(" [5] Access restrictions vocabulary:      " + str(Access_restrictions_vocabulary))
    print("[20] Contact point:                       " + str(Contact_point))
    print("[10] Publisher:                           " + str(Publisher))
    print("Rating Reusability:                       " + str(reusabilityPoints))
    print("---")
    print("Common rating:                            " + str(interoperabilityPoints + reusabilityPoints))
    print("---File---")
    
    printInfo(File_Info)



def makeExcel(fileName, url, Interoperability_Info, Reusability_Info, File_Info):
    if not File_Info:
        return
    
    fileName = fileName + ".xlsx"
    
    wb = openpyxl.Workbook()
    
    ws = wb[wb.sheetnames[0]]
    ws.title = "Лист1"
    
    label = [    
        "Format",
        "Media_type",
        "Format_Media_type_from_vocabulary",
        "Non_proprietary",
        "Machine_readable",
        "DCATAP_compliance",
        "License_information",
        "License_vocabulary",
        "Access_restrictions",
        "Access_restrictions_vocabulary",
        "Contact_point",
        "Publisher",
        "Rating",
        "Num_Rows",
        "Num_Columns"
    ]
    
    for i in range(1, 6 + 6 + 3 + 1):
        cell = ws.cell(row = 1, column = i)
        cell.value = label[i - 1]

    for i in range(1, 6 + 1):
        cell = ws.cell(row = 2, column = i)
        if Interoperability_Info[label[i - 1]]:
            cell.value = "+"
        else:
            cell.value = "-"
        
        cell = ws.cell(row = 2, column = i + 6)
        if Reusability_Info[label[i + 6 - 1]]:
            cell.value = "+"
        else:
            cell.value = "-"

    cell = ws.cell(row = 2, column = 13)
    cell.value = Interoperability_Info["InteroperabilityPoints"] + Reusability_Info["ReusabilityPoints"]

    num_columns = File_Info["num_columns"]

    cell = ws.cell(row = 2, column = 14)
    cell.value = File_Info["num_rows"]
    cell = ws.cell(row = 2, column = 15)
    cell.value = num_columns
    
    ws["B6"] = "пустые строки"
    ws["C6"] = "уникальные строки"
    ws["D6"] = "число 0"
    ws["E6"] = "максимум"
    ws["F6"] = "минимум"
    ws["G6"] = "среднее"
    
    column_names   = File_Info["column_names"]
    missing_values = File_Info["missing_values"]
    unique_values  = File_Info["unique_values"]
    amount_zero    = File_Info["amount_zero"]
    min_values     = File_Info["min_values"]
    max_values     = File_Info["max_values"]
    mean_values    = File_Info["mean_values"]
    
    for i in range(1, num_columns + 1):
        ws.cell(row = i + 6, column = 1).value = column_names[i - 1]
        ws.cell(row = i + 6, column = 2).value = missing_values[i - 1]
        ws.cell(row = i + 6, column = 3).value = unique_values[i - 1]
        ws.cell(row = i + 6, column = 4).value = amount_zero[i - 1]
        ws.cell(row = i + 6, column = 5).value = max_values[i - 1]
        ws.cell(row = i + 6, column = 6).value = min_values[i - 1]
        ws.cell(row = i + 6, column = 7).value = mean_values[i - 1]

    ws.cell(row = 6 + 3 + num_columns, column = 1).value = "ссылка"
    ws.cell(row = 6 + 3 + num_columns, column = 2).value = url

    wb.save(fileName)



# -----------------------------------------------



def checkOne(url, mediaTypeVocabulary, licencesVocabulary, generateExcelReport):
    r = requests.get(url)
    
    if r.status_code != 200:
        print("Ошибка:", r.status_code)
        return
    
    soup = bs(r.text, "html.parser")

    title = findTitle(soup)
    access = findAccessRestrictions(soup)
    license = findLicense(soup)
    formats = findFormats(soup)
    mediaTypes, mediaDownloadURL = findMediaType(soup)
    
    # Interoperability
    
    Format = haveFormats(formats)
    Media_type = haveMediaTypes(mediaTypes)
    Format_Media_type_from_vocabulary = isVocabularyMediaType(mediaTypes, mediaTypeVocabulary)
    
    Non_proprietary = isNonProprietaryFormat(formats)
    Machine_readable = isMachineReadableFormats(formats)
    
    DCATAP_compliance = checkComplianceDCATAP(mediaDownloadURL, createId(url))
    # DCATAP_compliance = False
    
    interoperabilityPoints = 0
    if Format:                            interoperabilityPoints +=  20
    if Media_type:                        interoperabilityPoints +=  10
    if Format_Media_type_from_vocabulary: interoperabilityPoints +=  10
    if Non_proprietary:                   interoperabilityPoints +=  20
    if Machine_readable:                  interoperabilityPoints +=  20
    if DCATAP_compliance:                 interoperabilityPoints +=  30
    
    Interoperability_Info = {
        "Format"                            : Format,
        "Media_type"                        : Media_type,
        "Format_Media_type_from_vocabulary" : Format_Media_type_from_vocabulary,
        "Non_proprietary"                   : Non_proprietary,
        "Machine_readable"                  : Machine_readable,
        "DCATAP_compliance"                 : DCATAP_compliance,
        "InteroperabilityPoints"            : interoperabilityPoints,
    }
    
    # Reusability
    
    License_information = haveLicense(license)
    License_vocabulary = isVocabularyLicense(license, licencesVocabulary)
    
    Access_restrictions = haveAccessRestrictions(access)
    Access_restrictions_vocabulary = isAccessRestrictionsVocabulary(access)
    
    Contact_point = haveContact(soup)
    Publisher = havePublisher(soup)
    
    reusabilityPoints = 0
    if License_information:            reusabilityPoints += 20
    if License_vocabulary:             reusabilityPoints += 10
    if Access_restrictions:            reusabilityPoints += 10
    if Access_restrictions_vocabulary: reusabilityPoints += 5
    if Contact_point:                  reusabilityPoints += 20
    if Publisher:                      reusabilityPoints += 10
    
    Reusability_Info = {
        "License_information"            : License_information,
        "License_vocabulary"             : License_vocabulary,
        "Access_restrictions"            : Access_restrictions,
        "Access_restrictions_vocabulary" : Access_restrictions_vocabulary,
        "Contact_point"                  : Contact_point,
        "Publisher"                      : Publisher,
        "ReusabilityPoints"              : reusabilityPoints,
    }
    
    # File
    
    fileDownloadURL = findDownloadLinks(soup)
    
    File_Info = checkFiles(fileDownloadURL, createId(url))
    
    # Output
    print(" @ title:       " + str(title))
    print(" @ formats:     " + str(formats))
    print(" @ media type:  " + str(mediaTypes))
    # print(" @ downloadURL: " + str(mediaDownloadURL))
    # print(" @ downloadURL: " + str(fileDownloadURL))
    print(" @ license:     " + str(license))
    print(" @ access:      " + str(access))
    
    printConsole(Interoperability_Info, Reusability_Info, File_Info)
    
    # Excel
    
    if generateExcelReport:
        if not os.path.exists("reports"):
            os.mkdir("reports")
        
        fileName = createId(url)
        filePath = os.path.join("reports", fileName)
        
        makeExcel(filePath, url, Interoperability_Info, Reusability_Info, File_Info)



def checkAll(listURL, generateExcelReport):
    mediaTypeVocabulary = getMediaTypeVocabulary()
    licencesVocabulary = getLicencesVocabulary()
    
    for index, url in enumerate(listURL):
        print("\n" + str(index + 1) + " --------------------------------------------\n")
        print(url)
        
        checkOne(url, mediaTypeVocabulary, licencesVocabulary, generateExcelReport)



# -----------------------------------------------



listURL = [
    "https://catalog.data.gov/dataset/death-rates-for-suicide-by-sex-race-hispanic-origin-and-age-united-states-020c1",
    "https://catalog.data.gov/dataset/drug-overdose-death-rates-by-drug-type-sex-age-race-and-hispanic-origin-united-states-3f72f",
    "https://catalog.data.gov/dataset/drug-use-data-from-selected-hospitals-26ee4",
    "https://catalog.data.gov/dataset/electric-vehicle-population-size-history-by-county",
    "https://catalog.data.gov/dataset/health-conditions-among-children-under-age-18-by-selected-characteristics-united-states-53b56",
    "https://catalog.data.gov/dataset/lottery-mega-millions-winning-numbers-beginning-2002",
    "https://catalog.data.gov/dataset/mental-health-care-in-the-last-4-weeks",
    "https://catalog.data.gov/dataset/meteorite-landings",
    "https://catalog.data.gov/dataset/nchs-leading-causes-of-death-united-states",
    "https://catalog.data.gov/dataset/school-attendance-by-student-group-and-district-2021-2022",
    "https://catalog.data.gov/dataset/street-names",
    "https://catalog.data.gov/dataset/2010-census-populations-by-zip-code",
    "https://catalog.data.gov/dataset/abs-store-inventory-and-sale-items",
    
    "https://catalog.data.gov/dataset/dataset-inventory",
    "https://catalog.data.gov/dataset/percent-of-covid-19-vaccine-recipients-who-live-in-a-svi-priority-zip-code-cumulative-and-",
    "https://catalog.data.gov/dataset/underground-storage-tanks-usts-facility-and-tank-details",
    "https://catalog.data.gov/dataset/post-covid-conditions-89bb3",
    "https://catalog.data.gov/dataset/u-s-life-expectancy-at-birth-by-state-and-census-tract-2010-2015",
    "https://catalog.data.gov/dataset/somerville-happiness-survey-responses",
    "https://catalog.data.gov/dataset/monthly-counts-of-deaths-by-select-causes-2014-2019-da9df",
    "https://catalog.data.gov/dataset/school-neighborhood-poverty-estimates-current",
    "https://catalog.data.gov/dataset/nchs-death-rates-and-life-expectancy-at-birth",
    "https://catalog.data.gov/dataset/nutrition-physical-activity-and-obesity-youth-risk-behavior-surveillance-system",
    "https://catalog.data.gov/dataset/borough-boundaries",
    "https://catalog.data.gov/dataset/supply-chain-shipment-pricing-data-07d29",
]

generateExcelReport = True



checkAll(listURL, generateExcelReport)

